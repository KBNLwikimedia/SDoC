"""
SDC ResolverURL + Captions Adder (for Wikimedia Commons)

----------

## Overview
Batch-updates Structured Data on Commons (SDC) for MediaInfo entities listed in
an Excel sheet. For each MediaInfo ID (M-id) it:

1. Ensures there is a statement **P7482 (source of file) = Q74228490 (file available on the internet)**,
   qualified with:
   * **P973 (described at URL)** = the row’s `ResolverURL`
   * **P137 (operator)** = **Q20670235 (Delpher)**

2. Adds **SDC captions** (labels) in **English (`en`)** and **Dutch (`nl`)**,
   derived from the **live Commons file title** (not the Excel, as these might contain erratic diacritics like
    De_TÄ³d_-_godsdienstig-staatkundig_dagblad_....), but **only if missing**:

   * Example live title: `File:Rotterdamsche courant 21-11-1840 (IA KBDDD02 000201168 mpeg21).pdf`
   * Derived caption candidate: `Rotterdamsche courant 21-11-1840`
     (removes `File:` prefix, common file extension, and the trailing ` (IA …)` segment)

Why two parts? The statement/qualifiers document that the file is available online at Delpher
via its persistent resolver, while captions improve search and display in SDC-enabled interfaces.

## Inputs
* **Excel file** (default: `"MediaFromDelpher-InternetArchiveFiles_16092025 - forSDC.xlsx"`)
* **Sheet** (default: `"MediaFromDelpher-IAfiles"`)
* **Required columns**:
  * `URL` — Commons MediaInfo entity link, e.g. `https://commons.wikimedia.org/entity/M109018409`
  * `ResolverURL` — e.g. `https://resolver.kb.nl/resolve?urn=KBDDD02:000201168`
* The Excel `Title` column is **not** used for captions (to avoid diacritic/typo drift).

## Outputs (Excel)
The same sheet is overwritten in place (sheet name preserved) with the following columns updated/added:
* `SDCStatus`            ∈ {`created`, `updated`, `already-present`, `error:<message>`}
* `CaptionENStatus`      ∈ {`created`, `already-present`, `skipped-empty-candidate`, `error:<message>`}
* `CaptionNLStatus`      ∈ {`created`, `already-present`, `skipped-empty-candidate`, `error:<message>`}
* `CaptionCandidate`     — the derived caption from the **live** title (for transparency/QA)

## Idempotence & Rules
* Statements (in sdc tab of Commons file):
  * If **no** P7482=Q74228490 exists → create one and add both qualifiers (P973, P137).
  * If it **exists** but is missing one/both qualifiers → add the missing qualifier(s).
  * If it already matches fully → `already-present`.
* Captions:
  * Reads existing labels via `wbgetentities`.
  * Sets **en**/**nl** only if missing; **does not overwrite** existing captions.
  * If the derived caption is empty (unexpected title pattern), both caption actions are skipped.

## APIs Used
All via the **Action API** (`/w/api.php`) and Special\\:EntityData:
* Login & tokens:
  * `action=query&meta=tokens&type=login`
  * `action=login` (BotPassword supported)
  * `action=query&meta=tokens&type=csrf`
* SDC statements:
  * Read:    `action=wbgetclaims&entity=M…&property=P7482`
  * Create:  `action=wbcreateclaim` (P7482=Q74228490)
  * Qualify: `action=wbsetqualifier` (P973=url, P137=Q20670235)
* Captions (labels) on MediaInfo:
  * Read: `action=wbgetentities&ids=M…&props=labels`
  * Set:  `action=wbsetlabel` (for `en` then `nl`)
* Live title resolution (correct diacritics):
  * M-id → pageid: `https://commons.wikimedia.org/wiki/Special:EntityData/M####.json`
  * pageid → title: `action=query&prop=info&pageids=<pageid>&formatversion=2`

## Authentication
Set the following environment variables (e.g., in a `.env` file):
* `WIKIMEDIA_USERNAME` — your username or BotPassword username (`User@BotName`)
* `WIKIMEDIA_PASSWORD` — your (Bot) password
* `WIKIMEDIA_USER_AGENT` — informative UA string per Wikimedia policy

## Configuration (env vars)
* `EXCEL_FILE`                (default: Excel file above)
* `EXCEL_SHEET`               (default: `MediaFromDelpher-IAfiles`)
* `EXCEL_URL_COL`             (default: `URL`)
* `EXCEL_RESOLVER_COL`        (default: `ResolverURL`)
* `EXCEL_STATUS_COL`          (default: `SDCStatus`)
* `EXCEL_CAPTION_EN_STATUS`   (default: `CaptionENStatus`)
* `EXCEL_CAPTION_NL_STATUS`   (default: `CaptionNLStatus`)
* `EXCEL_CAPTION_CANDIDATE`   (default: `CaptionCandidate`)
* Slicing: `HEAD` (first N rows) or `RANGE` (`N-M`, 1-based, inclusive) — **mutually exclusive**
* Politeness:
  * `EDIT_SLEEP_SEC`          (default: `0.2`) — delay between rows
  * Writes use `maxlag=5`
* Browser convenience:
  * `OPEN_AFTER_SUCCESS`      (1/0; open updated entity page in a new tab)
  * `OPEN_AFTER_SUCCESS_MAX`  (cap number of tabs)

## Dependencies
* Python packages: `requests`, `pandas`, `openpyxl`, `python-dotenv`, `tqdm`
* Internet access to `commons.wikimedia.org`

## Error Handling & Edge Cases
* MediaInfo with **no SDC**: Action API may respond with `no-such-entity`; this is treated as “no claims/labels yet”.
* Missing/invalid `ResolverURL`: statement part logs an error but captions still proceed.
* Missing `pageid` or deleted/moved file: captions skipped with an explanatory error.
* Derived caption empty/unusable: captions skipped (`skipped-empty-candidate`).
* All API errors are logged with context; per-row status records success/skip/error without halting the entire batch.

## Performance & Courtesy
* Uses `maxlag=5` on mutating calls and a small sleep between rows (configurable).
* Includes an explicit `User-Agent`.
* Suitable for batch runs; use `HEAD`/`RANGE` to test on a subset first.

## Safety Notes
* The script **does not overwrite existing captions**.
* Statement updates are narrowly scoped to P7482/Q74228490 and the two qualifiers.
* Keep your BotPassword credentials secure (e.g., via `.env`, not committed to VCS).

------------------

Author: ChatGPT, prompted by Olaf Janssen, Wikimedia coordinator at KB, Koninklijke Bibliotheek, national library of the Netherlands
Latest update: 2025-10-08
License: CC0
"""
# ====================================================================
# ----------- Imports
# ====================================================================

import json
import logging
import os
import random
import re
import time
import webbrowser
from typing import Dict, List, Optional, Tuple, Any
from urllib.parse import urlparse

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from tqdm.auto import tqdm

# ====================================================================
# ---------------- Configuration ----------------
# ====================================================================
load_dotenv()

API_URL = "https://commons.wikimedia.org/w/api.php"

EXCEL_FILE = os.getenv("EXCEL_FILE", "MediaFromDelpher-InternetArchiveFiles_16092025 - forSDC.xlsx")
EXCEL_SHEET_ENV = os.getenv("EXCEL_SHEET", "").strip()
EXCEL_SHEET = "MediaFromDelpher-IAfiles" if not EXCEL_SHEET_ENV else EXCEL_SHEET_ENV

EXCEL_URL_COL = os.getenv("EXCEL_URL_COL", "URL")
EXCEL_RESOLVER_COL = os.getenv("EXCEL_RESOLVER_COL", "ResolverURL")

EXCEL_STATUS_COL = os.getenv("EXCEL_STATUS_COL", "SDCStatus")
EXCEL_CAPTION_EN_STATUS = os.getenv("EXCEL_CAPTION_EN_STATUS", "CaptionENStatus")
EXCEL_CAPTION_NL_STATUS = os.getenv("EXCEL_CAPTION_NL_STATUS", "CaptionNLStatus")
EXCEL_CAPTION_CANDIDATE = os.getenv("EXCEL_CAPTION_CANDIDATE", "CaptionCandidate")

HEAD = int(os.getenv("HEAD", "0"))         # if >0, use first HEAD rows
RANGE = os.getenv("RANGE", "1-100").strip()     # "N-M" (1-based inclusive, e.g. 1-5 for Excel rows 2 t/m 6)

SLEEP_BETWEEN = float(os.getenv("EDIT_SLEEP_SEC", "4.0"))  # polite delay between rows
OPEN_AFTER_SUCCESS = int(os.getenv("OPEN_AFTER_SUCCESS", "0")) # yes or no to open updated entity in browser
OPEN_AFTER_SUCCESS_MAX = int(os.getenv("OPEN_AFTER_SUCCESS_MAX", "5")) # cap number of browser tabs

RETRYABLE_STATUS = {502, 503, 504}
RETRYABLE_API_CODES = {"maxlag", "ratelimited", "internal_api_error"}
RETRYABLE_AUTH_CODES = {"badtoken", "assertuserfailed", "assertnameduserfailed", "notloggedin"}

# Checkpoint (flush) to Excel after this many successful edits (default 50)
CHECKPOINT_EVERY_SUCCESS = int(os.getenv("CHECKPOINT_EVERY_SUCCESS", "50"))

USERNAME = os.getenv("WIKIMEDIA_USERNAME", "").strip()
PASSWORD = os.getenv("WIKIMEDIA_PASSWORD", "").strip()
USER_AGENT = os.getenv("WIKIMEDIA_USER_AGENT", "KB-SDC-ResolverURL-Captions/1.0 (olaf.janssen@kb.nl) - Olaf Janssen, KB, national library of the Netherlands").strip()

# SDC constants
P_SOURCE_OF_FILE = "P7482"        # source of file
Q_FILE_ONLINE = 74228490          # Q74228490 file available on the internet
P_DESCRIBED_AT_URL = "P973"       # described at URL
P_OPERATOR = "P137"               # operator
Q_DELPHER = 20670235              # Q20670235 Delpher

MID_RE = re.compile(r"/(M\d+)(?:[/?#].*)?$", re.IGNORECASE)

# Strip trailing " (IA ...).pdf" (case-insensitive) from live title (after removing "File:")
IA_TRAIL_RE = re.compile(r"\s*\(IA[^)]*\)\s*$", re.IGNORECASE)

# ====================================================================
#  Class and functions: Logging & progress-bar integration
# This group wires Python’s logging to play nicely with tqdm so you
# get one clean log line per row while the progress bar keeps rendering correctly.
# ====================================================================

class TqdmLoggingHandler(logging.Handler):
    """
    A logging handler that plays nicely with `tqdm` progress bars.

    Why
    ----
    Normal `logging` writes directly to the console stream while `tqdm` redraws
    the same line for its progress bar. Interleaving the two can corrupt the bar
    display. This handler routes log lines through `tqdm.write(...)`, which
    prints on a clean line above the bar without breaking it.

    Behavior
    --------
    - Formats each `LogRecord` with the handler’s formatter.
    - Emits via `tqdm.write(...)` when `tqdm` is available.
    - Forces a refresh on all active bars (`tqdm._instances`) so that the
      progress display stays responsive even when `tqdm`’s `mininterval`
      throttling would otherwise delay a redraw.
    - If `tqdm` is not importable at runtime, falls back to writing to `sys.stderr`
      so logs are still visible.

    Notes
    -----
    - `tqdm._instances` is an internal attribute; this handler guards its use
      with `getattr` and individual `try/except` blocks for safety.
    - Install this handler on the root logger (or your app logger) and remove
      other console handlers to avoid duplicate lines.
    - Thread-friendly: `tqdm.write` uses an internal lock.

    Example
    -------
        import logging
        from tqdm import tqdm

        logger = logging.getLogger()
        logger.handlers.clear()
        h = TqdmLoggingHandler()
        h.setFormatter(logging.Formatter("%(asctime)s - %(levelname)s - %(message)s"))
        logger.addHandler(h)
        logger.setLevel(logging.INFO)

        for i in tqdm(range(100), desc="Work", unit="it"):
            if i % 10 == 0:
                logger.info("Reached %s", i)

    """

    def emit(self, record: logging.LogRecord) -> None:
        try:
            msg = self.format(record)

            # Try to emit via tqdm so we don't break the bar.
            try:
                from tqdm import tqdm as _tqdm
            except Exception:
                # Fallback: no tqdm available — write to stderr.
                import sys
                sys.stderr.write(msg + "\n")
                sys.stderr.flush()
                return

            _tqdm.write(msg)

            # Force-refresh all active bars (defensive; _instances is internal).
            instances = getattr(_tqdm, "_instances", None)
            if instances:
                for bar in list(instances):
                    try:
                        bar.refresh()
                    except Exception:
                        # Never let a single bar’s failure break logging.
                        pass

        except Exception:
            # Delegate to logging’s standard error handling.
            self.handleError(record)


def setup_tqdm_logging(level=logging.INFO):
    """
    Install a console logging handler that plays nicely with `tqdm` progress bars.

    Why
    ----
    Standard `logging` writes directly to the stream while `tqdm` continually
    rewrites the same line to render the progress bar. Mixing the two can cause
    garbled output or bars that "jump". This function replaces the root logger’s
    console handler with a `TqdmLoggingHandler` that routes messages through
    `tqdm.write(...)`, ensuring each log line is printed cleanly on its own line
    without breaking the bar.

    Behavior
    --------
    - Resolves and applies the requested logging level (accepts `int` or `"INFO"`, etc.).
    - Removes existing handlers from the root logger to prevent duplicate lines.
    - Adds a single `TqdmLoggingHandler` if available; otherwise falls back to a
      normal `StreamHandler` so logging still works.
    - Sets a timestamped, level-prefixed format.
    - Sets `logger.propagate = False` to avoid double-printing via ancestor loggers.
    - Safe to call multiple times; it re-installs a single console handler.

    Parameters
    ----------
    level : int | str
        Logging verbosity (e.g., `logging.INFO`, `logging.DEBUG`, or the string
        names `"INFO"`, `"DEBUG"`, ...). Defaults to `logging.INFO`.

    Raises
    ------
    ValueError
        If `level` cannot be resolved to a valid numeric logging level.

    Notes
    -----
    - This affects the *root* logger. If you have per-module loggers with their
      own handlers, they may still emit unless you also adjust them.
    - Keep the format relatively short so the progress bar remains readable.
    """
    # Resolve level if given as a string (e.g., "INFO")
    if isinstance(level, str):
        resolved = getattr(logging, level.upper(), None)
        if not isinstance(resolved, int):
            raise ValueError(f"Invalid logging level string: {level!r}")
        level = resolved
    if not isinstance(level, int):
        raise ValueError(f"Invalid logging level type: {type(level).__name__}")

    logger = logging.getLogger()
    logger.setLevel(level)

    # Remove any existing handlers to avoid double printing
    try:
        logger.handlers.clear()
    except Exception:
        # Older Python versions: fallback to manual removal
        for h in list(logger.handlers):
            logger.removeHandler(h)

    # Install a tqdm-aware handler; fall back to a plain stream handler if needed
    try:
        h = TqdmLoggingHandler()  # assumes class is defined/imported elsewhere
    except Exception:
        h = logging.StreamHandler()

    h.setLevel(level)
    # Leading " - " visually separates log lines from the tqdm bar
    h.setFormatter(logging.Formatter(" - %(asctime)s - %(levelname)s - %(message)s"))
    logger.addHandler(h)

    # Prevent propagation to ancestor loggers (avoids duplicate output in some IDEs)
    logger.propagate = False


# ====================================================================
#  Functions: Excel I/O & workload selection
# These functions read/shape the workload from Excel (supporting HEAD/RANGE) and
# persist status columns back to the same sheet, including periodic checkpoints to avoid data
# loss on long runs.
# ====================================================================

def apply_slice(df: pd.DataFrame) -> pd.DataFrame:
    """
    Return a sliced view of the input DataFrame based on global RANGE or HEAD.

    Rules (strict):
      - Exactly one of RANGE or HEAD may be set; if both are set ⇒ abort.
      - RANGE must be a 1-based inclusive span "N-M" (e.g., "5-12").
        * Validates format and bounds: 1 <= N <= M <= len(df).
        * Returns rows df.iloc[N-1:M].copy()
      - HEAD must be an integer with 1 <= HEAD <= len(df).
        * Returns df.head(HEAD).copy()
      - If neither is set (RANGE empty/blank and HEAD <= 0), returns the original df unchanged.

    Args:
        df: Source pandas DataFrame (must be a DataFrame).
    Returns:
        pandas.DataFrame: The sliced DataFrame (copy) or the original df if no slicing.
    Raises:
        SystemExit: On invalid configuration (both set), bad RANGE format, or out-of-bounds values.
    Notes:
        - This function does not mutate the input DataFrame.
        - Logging emits INFO messages indicating the slice used.
    """
    # Validate input
    if not isinstance(df, pd.DataFrame):
        logging.critical("apply_slice: df must be a pandas DataFrame, got %r", type(df).__name__)
        raise SystemExit(1)

    if df.empty:
        logging.info("apply_slice: input DataFrame is empty; nothing to slice.")
        return df

    # Normalize config flags
    rng = str(RANGE).strip() if isinstance(RANGE, str) else (str(RANGE).strip() if RANGE is not None else "")
    has_range = bool(rng)
    try:
        head_val = int(HEAD)
    except Exception:
        head_val = 0
    has_head = head_val > 0

    # Mutually exclusive
    if has_range and has_head:
        logging.critical("apply_slice: Both RANGE and HEAD are set; choose one.")
        raise SystemExit(1)

    nrows = len(df)

    if has_range:
        m = re.match(r"^\s*(\d+)\s*[-:]\s*(\d+)\s*$", rng)
        if not m:
            logging.critical("apply_slice: RANGE must be 'N-M'. Got: %r", RANGE)
            raise SystemExit(1)
        start, end = int(m.group(1)), int(m.group(2))
        if not (1 <= start <= end <= nrows):
            logging.critical("apply_slice: RANGE out of bounds for %d rows: %s", nrows, RANGE)
            raise SystemExit(1)
        logging.info("apply_slice: Using RANGE %d-%d (1-based inclusive) out of %d.", start, end, nrows)
        return df.iloc[start - 1:end].copy()

    if has_head:
        if not (1 <= head_val <= nrows):
            logging.critical("apply_slice: HEAD must be 1..%d, got %d", nrows, head_val)
            raise SystemExit(1)
        logging.info("apply_slice: Using HEAD=%d out of %d.", head_val, nrows)
        return df.head(head_val).copy()

    logging.info("apply_slice: No RANGE/HEAD set; processing all %d rows.", nrows)
    return df


def write_back(df_full: pd.DataFrame, target_sheet: str) -> None:
    """
    Writes the in-memory DataFrame to the Excel workbook.

    Behavior:
        - Replaces the sheet named `target_sheet` in `EXCEL_FILE` if it exists,
          otherwise creates it.
        - Uses the openpyxl engine and writes without the DataFrame index.
        - Falls back to create the workbook (mode='w') if `EXCEL_FILE` does not exist.
    Args:
        df_full: The pandas DataFrame to write.
        target_sheet: Name of the worksheet to replace/create (non-empty string).
    Raises:
        SystemExit: If validation fails or the workbook cannot be written.
                    (Common cause: the file is open/locked in Excel.)
    Notes:
        - If you see a PermissionError, close the Excel file and re-run.
        - For very large sheets consider enabling write-only mode or chunked writes.
    """
    # Validate inputs
    if not isinstance(df_full, pd.DataFrame):
        logging.critical("write_back: df_full must be a pandas DataFrame, got %r", type(df_full).__name__)
        raise SystemExit(1)
    if not isinstance(target_sheet, str) or not target_sheet.strip():
        logging.critical("write_back: target_sheet must be a non-empty string, got %r", target_sheet)
        raise SystemExit(1)

    mode = "a" if os.path.exists(EXCEL_FILE) else "w"
    try:
        with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode=mode, if_sheet_exists="replace") as w:
            df_full.to_excel(w, sheet_name=target_sheet, index=False)
        logging.info("Excel updated: %s (sheet=%r)", EXCEL_FILE, target_sheet)
    except PermissionError as e:
        logging.critical("write_back: permission error writing %s (is it open in Excel?): %s", EXCEL_FILE, e)
        raise SystemExit(1)
    except Exception as e:
        logging.critical("write_back: failed to write %s (sheet=%r): %s", EXCEL_FILE, target_sheet, e)
        raise SystemExit(1)

# ====================================================================
#  Functions: URL & title utilities
# Small helpers for parsing and presentation: extract the MediaInfo ID from a URL,
# derive the pageid from the M-id (no network), generate clean caption
# candidates from live titles, and optionally open the entity page in a browser.
# ====================================================================
def extract_mid(entity_url: str) -> Optional[str]:
    """
    Extract a Commons MediaInfo ID (e.g., "M109018409") from a string.

    Accepts:
      - Full entity URLs such as "https://commons.wikimedia.org/entity/M109018409"
      - Strings that already are a bare MID (e.g., "M109018409" or "m109018409")
    Behavior:
      - Returns the normalized MID (uppercase "M" + digits) on success.
      - Returns None on invalid/empty input or when no MID can be found.
      - Never raises; unexpected parsing issues are logged at WARNING level.
    Args:
        entity_url: URL or identifier to parse.
    Returns:
        Optional[str]: The MediaInfo ID ("M" + digits) or None if not found.
    """
    try:
        if not isinstance(entity_url, str):
            return None
        s = entity_url.strip()
        if not s:
            return None

        # Try URL-style match first (…/M####…)
        m = MID_RE.search(s)
        if m:
            return m.group(1).upper()

        # Also accept bare IDs like "M12345" or "m12345"
        if re.fullmatch(r"[Mm]\d+", s):
            return s.upper()

        logging.debug("extract_mid: no MID found in %r", entity_url)
        return None
    except Exception as e:
        logging.warning("extract_mid: failed to parse %r: %s", entity_url, e)
        return None


def pageid_from_mid(mid: str) -> Optional[int]:
    """
    Parse a Commons MediaInfo ID into its underlying File page ID.

    On Wikimedia Commons, a MediaInfo ID is simply the File page ID prefixed
    with the letter 'M' (case-insensitive), e.g.:
        - "M109018409"  →  109018409
        - "m42"         →  42
    Behavior:
        - Strips surrounding whitespace.
        - Accepts only the exact pattern `[Mm]<digits>` (no extra suffixes/prefixes).
        - Returns the integer page ID on success; otherwise `None`.
        - Silently returns `None` for non-string inputs or malformed IDs to keep
          callers' control-flow simple; no exception is raised for format issues.
    Args:
        mid: The MediaInfo identifier string (e.g., "M12345").
    Returns:
        Optional[int]: The numeric page ID if parsing succeeds; otherwise `None`.
    Notes:
        - This function performs no network requests.
        - If you prefer strict failure for bad input, validate `mid` before calling.
    """
    if not isinstance(mid, str):
        return None
    s = mid.strip()
    if not s:
        return None
    m = re.fullmatch(r"[Mm](\d+)", s)
    if not m:
        return None
    try:
        return int(m.group(1))
    except (ValueError, OverflowError):
        # Extremely large or otherwise invalid numeric text; treat as unusable.
        logging.warning("pageid_from_mid: numeric part not an int: %r", m.group(1))
        return None


def derive_caption_from_live_title(live_title: str) -> str:
    """
    Return a clean caption candidate derived from a live Commons File title.

    Transformations (in order):
      1) Remove a leading ``"File:"`` prefix (case-insensitive).
      2) Replace underscores with spaces.
      3) Strip a *known* file extension at the very end (``.pdf``, ``.jpg``,
         ``.jpeg``, ``.png``, ``.tif``, ``.tiff``, ``.djvu``; case-insensitive).
      4) Remove a trailing ``" (IA ...)"`` parenthetical (case-insensitive),
         e.g. ``" (IA KBDDD02 000201168 mpeg21)"``.
      5) Trim surrounding whitespace and collapse internal runs of whitespace.
    Robustness:
      - Never raises; returns ``""`` on non-string/empty input.
      - Catches unexpected errors, logs a warning, and falls back to a trimmed
        version of the original title.
    Args:
        live_title: The canonical page title returned by the API, e.g.
            ``"File:Rotterdamsche courant 21-11-1840 (IA KBDDD02 000201168 mpeg21).pdf"``.
    Returns:
        A human-readable caption candidate, e.g. ``"Rotterdamsche courant 21-11-1840"``,
        or ``""`` if no reasonable candidate can be produced.

    Examples:
        >>> derive_caption_from_live_title("File:Foo_bar (IA ddd 012345678).pdf")
        'Foo bar'
        >>> derive_caption_from_live_title("File:Image name.TIFF")
        'Image name'
        >>> derive_caption_from_live_title(None)
        ''
    """

    if not isinstance(live_title, str) or not live_title.strip():
        return ""

    try:
        name = live_title.strip()

        # 1) strip "File:" prefix
        if name[:5].lower() == "file:":
            name = name[5:]

        # 2) underscores -> spaces
        name = name.replace("_", " ").strip()

        # 3) remove known file extension at end
        m = re.search(r"\.(pdf|jpg|jpeg|png|tif|tiff|djvu)$", name, re.IGNORECASE)
        if m:
            name = name[: -len(m.group(0))]

        # 4) remove trailing " (IA ...)"
        # assumes IA_TRAIL_RE = re.compile(r"\s*\(IA[^)]*\)\s*$", re.IGNORECASE)
        try:
            name = IA_TRAIL_RE.sub("", name).strip()
        except NameError:
            # If the regex isn't available, skip this step silently.
            pass

        # 5) collapse internal whitespace
        name = re.sub(r"\s+", " ", name).strip()

        return name
    except Exception as e:
        logging.warning("derive_caption_from_live_title: unexpected error for title %r: %s", live_title, e)
        try:
            return str(live_title).strip()
        except Exception:
            return ""

def open_page_in_browser(mid: str, title_hint: Optional[str]) -> None:
    """
    Open the Commons MediaInfo entity page in a new browser tab.

    This constructs and opens:
        https://commons.wikimedia.org/wiki/Special:EntityPage/{mid}

    Using Special:EntityPage avoids title-encoding issues and works
    regardless of the linked File page’s exact title or later renames.
    Args:
        mid: MediaInfo ID (e.g., "M109018409"). Case-insensitive; must be a non-empty string.
        title_hint: Optional human-readable title shown only in the log message.
    Behavior:
        - Logs the target URL (and the title hint, if provided).
        - Attempts to open a new tab via Python's `webbrowser` module.
        - If `mid` is empty/invalid, logs a WARNING and returns without raising.
        - Any exception from the browser invocation is caught and logged as a WARNING.
          The function never raises.
    Notes:
        - `webbrowser.open_new_tab()` is best-effort and platform-dependent; it may
          return without focusing the browser (or do nothing in headless environments).
        - This helper is typically called after a successful change so you can quickly
          inspect the updated entity.
    """
    try:
        if not isinstance(mid, str) or not mid.strip():
            logging.warning("open_page_in_browser: invalid MID: %r", mid)
            return
        url = f"https://commons.wikimedia.org/wiki/Special:EntityPage/{mid}"
        logging.info("Opening in browser: %s%s", url, f" (Title≈{title_hint!r})" if title_hint else "")
        webbrowser.open_new_tab(url)
    except Exception as e:
        logging.warning("Failed to open browser for %s: %s", mid, e)



# ====================================================================
#  Functions: MediaWiki / HTTP primitives
# Low-level client utilities used everywhere else: a thin GET/POST wrapper, login + CSRF retrieval,
# a CSRF refresh helper, and a robust mutate-with-retries helper that
# handles maxlag, badtoken, transient 5xx, and Retry-After.
# ====================================================================

def mw_request(session: requests.Session, method: str, params: Dict, data: Dict = None) -> Dict:
    """
    Perform a single MediaWiki Action API request with robust error reporting.

    This thin wrapper standardizes headers, validates the HTTP method, applies sane timeouts,
    and logs helpful diagnostics (with sensitive fields redacted) when requests fail.

    Args:
        session: A configured `requests.Session` (cookies/auth already established).
        method:  "GET" or "POST" (case-insensitive).
        params:  Querystring parameters sent to the Action API.
        data:    Form data for POST requests (ignored for GET). Optional.

    Returns:
        dict: The parsed JSON response.

    Raises:
        ValueError:              If `method` is not GET or POST.
        requests.Timeout:        On request timeout.
        requests.HTTPError:      For non-2xx responses (after `raise_for_status()`).
        requests.RequestException:
                                 For other network/transport errors.
        ValueError (JSON decode):
                                 If the response body is not valid JSON.
    Notes:
        - Sensitive fields (e.g., tokens/passwords) are redacted in logs.
        - Timeouts: 30s for GET, 60s for POST.
    """


    def _redact(d: Optional[Dict]) -> Optional[Dict]:
        if not isinstance(d, dict):
            return d
        REDACT = {"token", "csrftoken", "logintoken", "lgtoken", "lgpassword", "password", "assertuser", "Authorization"}
        masked = {}
        for k, v in d.items():
            if str(k).lower() in {s.lower() for s in REDACT}:
                masked[k] = "<redacted>"
            else:
                masked[k] = v
        return masked

    method_upper = (method or "").upper()
    if method_upper not in {"GET", "POST"}:
        raise ValueError(f"mw_request: unsupported method {method!r}; expected 'GET' or 'POST'.")

    headers = {"User-Agent": USER_AGENT}
    try:
        if method_upper == "GET":
            r = session.get(API_URL, params=params, headers=headers, timeout=30)
        else:
            r = session.post(API_URL, params=params, data=(data or {}), headers=headers, timeout=60)

        r.raise_for_status()

        try:
            return r.json()
        except ValueError as e:
            preview = r.text[:300] if hasattr(r, "text") else "<no-body>"
            logging.error(
                "mw_request: JSON parse error for %s %s | params=%s data=%s | body-start=%r",
                method_upper, API_URL, _redact(params), _redact(data), preview
            )
            raise

    except requests.Timeout as e:
        logging.error(
            "mw_request: timeout for %s %s | params=%s data=%s",
            method_upper, API_URL, _redact(params), _redact(data)
        )
        raise
    except requests.HTTPError as e:
        status = getattr(e.response, "status_code", "?")
        body_preview = ""
        try:
            body_preview = e.response.text[:300]
        except Exception:
            body_preview = "<no-body>"
        logging.error(
            "mw_request: HTTP %s for %s %s | params=%s data=%s | body-start=%r",
            status, method_upper, API_URL, _redact(params), _redact(data), body_preview
        )
        raise
    except requests.RequestException as e:
        logging.error(
            "mw_request: request error for %s %s | params=%s data=%s | err=%s",
            method_upper, API_URL, _redact(params), _redact(data), e
        )
        raise


def mw_login_and_tokens() -> Tuple[requests.Session, str]:
    """
    Create an authenticated MediaWiki session and fetch a CSRF token.

    Workflow:
      1) Validate that `WIKIMEDIA_USERNAME` and `WIKIMEDIA_PASSWORD` are present.
      2) GET a **login token** via:
           action=query&meta=tokens&type=login
      3) POST **action=login** with username, password, and the login token.
         (Works with normal accounts or BotPasswords; use `User@BotName` for the latter.)
      4) GET a **CSRF token** via:
           action=query&meta=tokens&type=csrf
      5) Return the (session, csrf_token) pair.

    Returns:
        (requests.Session, str): An authenticated session (with cookies) and a non-empty CSRF token.

    Strict failure behavior:
        - Logs a CRITICAL message and terminates the run with SystemExit(1) on any error
          (missing env vars, HTTP error, non-success login, or missing tokens).

    Notes:
        - All network calls are made through `mw_request`, which sets the User-Agent and raises
          on non-2xx responses. This function wraps those failures and aborts cleanly.
        - Sensitive values (passwords/tokens) are **not** logged.
    """
    if not USERNAME or not PASSWORD:
        logging.critical("Missing WIKIMEDIA_USERNAME or WIKIMEDIA_PASSWORD in environment.")
        raise SystemExit(1)

    session = requests.Session()

    # --- Step 1: login token ---
    try:
        tok = mw_request(session, "GET", {
            "action": "query",
            "meta": "tokens",
            "type": "login",
            "format": "json"
        })
    except Exception as e:
        logging.critical("Failed to obtain login token: %s", e)
        raise SystemExit(1)

    login_token = tok.get("query", {}).get("tokens", {}).get("logintoken")
    if not login_token:
        logging.critical("Failed to obtain login token (empty response).")
        raise SystemExit(1)

    # --- Step 2: login ---
    try:
        login = mw_request(session, "POST",
            {"action": "login", "format": "json"},
            data={"lgname": USERNAME, "lgpassword": PASSWORD, "lgtoken": login_token}
        )
    except Exception as e:
        logging.critical("Login request failed: %s", e)
        raise SystemExit(1)

    login_result = (login.get("login") or {}).get("result")
    if login_result != "Success":
        # Avoid printing sensitive details; include MW code if present
        code = (login.get("login") or {}).get("reason") or login_result or "unknown"
        logging.critical("Login failed (result=%r).", code)
        raise SystemExit(1)

    # --- Step 3: csrf token ---
    try:
        tok2 = mw_request(session, "GET", {
            "action": "query",
            "meta": "tokens",
            "type": "csrf",
            "format": "json"
        })
    except Exception as e:
        logging.critical("Failed to obtain CSRF token: %s", e)
        raise SystemExit(1)

    csrf = tok2.get("query", {}).get("tokens", {}).get("csrftoken")
    if not csrf:
        logging.critical("Failed to obtain CSRF token (empty response).")
        raise SystemExit(1)

    return session, csrf


def fetch_csrf(session: requests.Session, api_url: str, user_agent: str) -> str:
    """
    Fetch a fresh CSRF (edit) token from a MediaWiki API.

    Args:
        session (requests.Session): An authenticated session (after a successful `action=login`).
        api_url (str): The API endpoint URL, e.g. "https://commons.wikimedia.org/w/api.php".
        user_agent (str): User-Agent string to send per Wikimedia API policy.

    Returns:
        str: A non-empty CSRF token suitable for write actions (e.g., wbcreateclaim, wbsetlabel).

    Raises:
        TypeError: If input types/values are invalid.
        RuntimeError: If the HTTP request fails, the response is not JSON, or the token
                      is missing/empty in the response.

    Notes:
        - CSRF tokens can expire; on "badtoken" errors, call this again to refresh.
        - This function does NOT (re)login; it assumes `session` already holds valid auth cookies.
    """
    # Basic validation
    if not isinstance(session, requests.Session):
        raise TypeError("fetch_csrf: 'session' must be a requests.Session")
    if not isinstance(api_url, str) or not api_url.strip():
        raise TypeError("fetch_csrf: 'api_url' must be a non-empty string")
    if not isinstance(user_agent, str) or not user_agent.strip():
        raise TypeError("fetch_csrf: 'user_agent' must be a non-empty string")

    try:
        r = session.get(
            api_url,
            params={"action": "query", "meta": "tokens", "type": "csrf", "format": "json"},
            headers={"User-Agent": user_agent},
            timeout=30,
        )
        r.raise_for_status()
        try:
            j = r.json()
        except ValueError as e:
            raise RuntimeError(f"CSRF token response is not valid JSON: {e}") from e

        token = (j.get("query") or {}).get("tokens", {}).get("csrftoken")
        if not token or not isinstance(token, str) or not token.strip():
            raise RuntimeError("No CSRF token found in API response")
        return token
    except requests.RequestException as e:
        # Network/HTTP-layer errors (timeouts, 4xx/5xx after raise_for_status, etc.)
        raise RuntimeError(f"Failed to fetch CSRF token: {e}") from e


def mw_mutate(session: requests.Session,
              csrf_token: str,
              params: dict,
              data: dict = None,
              *,
              api_url: str,
              user_agent: str,
              max_retries: int = 4,
              base_sleep: float = 1.0) -> tuple[Any, str] | None:
    """
    Perform a **POST** to the MediaWiki Action API with polite retries, backoff, and optional CSRF refresh.

    This helper centralizes write calls (e.g., `wbcreateclaim`, `wbsetqualifier`, `wbsetlabel`) and
    takes care of common concerns:

    - Injects stable parameters on every attempt:
      `format=json`, `maxlag=5` (in **params**), and in **data**: `assert=user`, `bot=1`, `token=<csrf>`.
    - Retries on transient HTTP errors (e.g., 429/502/503/504) honoring the `Retry-After` header when present,
      otherwise using exponential backoff with jitter.
    - Retries on common API-layer transient conditions (e.g., `maxlag`, `ratelimited`, `internal_api_error_*`).
    - Refreshes the CSRF token **once per failure sequence** when the API reports token/auth issues
      (e.g., `badtoken`, `notloggedin`, `assertuserfailed`), by calling `fetch_csrf(session, api_url, user_agent)`.
    - Returns the parsed JSON payload and the **CSRF token actually used** (which may be refreshed).

    Args:
        session (requests.Session): An authenticated session (after successful login).
        csrf_token (str): A current CSRF token to start with; may be refreshed internally.
        params (dict): Querystring parameters for the action (e.g., `{"action": "wbsetlabel"}`).
        data (dict, optional): POST form data for the action. Defaults to `{}` if omitted.
        api_url (str): MediaWiki API endpoint, e.g. `"https://commons.wikimedia.org/w/api.php"`.
        user_agent (str): User-Agent string to send per Wikimedia API policy.
        max_retries (int): Max number of retries for transient failures (default: 4).
        base_sleep (float): Base seconds used for backoff (default: 1.0).

    Returns:
        tuple[Any, str] | None:
            `(json_response, csrf_token_used_or_refreshed)` on success.
            The function never returns `None` on success; `None` appears only if the caller
            annotates a wider union and swallows exceptions elsewhere.

    Raises:
        TypeError: If required arguments have invalid types/values.
        RuntimeError: For non-retryable API errors (a JSON `"error"` payload not in the retry lists).
        requests.RequestException: When HTTP/transport ultimately fails after retries.

    Notes:
        - This function **does not** (re)login. If the session cookies are invalid, refresh CSRF may still fail.
        - `maxlag=5` lets the server defer your write when replicas lag; we retry with backoff automatically.
        - Keep your `user_agent` informative; Wikimedia requests this for responsible bot operation.

    Example:
        * params = {"action": "wbsetlabel"}
        * data = {"id": "M123", "language": "en", "value": "A caption"}
        * resp, csrf = mw_mutate(sess, csrf, params, data, api_url=API_URL, user_agent=UA)
        * assert "success" in resp.get("edit", {}) or "entity" in resp
    """
    # --- Basic input validation (fast-fail with clear errors) -----------------
    if not isinstance(session, requests.Session):
        raise TypeError("mw_mutate: 'session' must be a requests.Session")
    if not isinstance(csrf_token, str) or not csrf_token:
        raise TypeError("mw_mutate: 'csrf_token' must be a non-empty string")
    if not isinstance(params, dict):
        raise TypeError("mw_mutate: 'params' must be a dict")
    if data is not None and not isinstance(data, dict):
        raise TypeError("mw_mutate: 'data' must be a dict or None")
    if not isinstance(api_url, str) or not api_url.strip():
        raise TypeError("mw_mutate: 'api_url' must be a non-empty string")
    if not isinstance(user_agent, str) or not user_agent.strip():
        raise TypeError("mw_mutate: 'user_agent' must be a non-empty string")
    if not isinstance(max_retries, int) or max_retries < 0:
        raise TypeError("mw_mutate: 'max_retries' must be an int ≥ 0")
    if not isinstance(base_sleep, (int, float)) or base_sleep <= 0:
        raise TypeError("mw_mutate: 'base_sleep' must be a positive number")

    # Make shallow copies so we can mutate safely
    p = dict(params) if params else {}
    d = dict(data) if data else {}

    # stable params
    p.setdefault("format", "json")
    p.setdefault("maxlag", "5")

    attempt = 0
    token = csrf_token

    while True:
        attempt += 1

        # inject auth-ish fields each attempt (token may refresh)
        d["token"] = token
        d.setdefault("assert", "user")
        d.setdefault("bot", "1")

        try:
            r = session.post(api_url, params=p, data=d,
                             headers={"User-Agent": user_agent}, timeout=60)
            # Respect Retry-After on 429/5xx, etc.
            if r.status_code in RETRYABLE_STATUS:
                ra = r.headers.get("Retry-After")
                if ra is not None:
                    try:
                        sleep = max(float(ra), base_sleep)
                    except Exception:
                        sleep = base_sleep
                else:
                    # exponential backoff with jitter
                    sleep = base_sleep * (2 ** (attempt - 1)) * (0.8 + 0.4 * random.random())
                if attempt <= max_retries:
                    logging.warning("API %s: HTTP %s, retrying in %.1fs (attempt %d/%d)",
                                    p.get("action"), r.status_code, sleep, attempt, max_retries)
                    time.sleep(sleep)
                    continue
                r.raise_for_status()

            r.raise_for_status()
            j = r.json()

            # Handle API-level errors inside JSON
            if "error" in j:
                code = (j["error"].get("code") or "").lower()
                if code in RETRYABLE_API_CODES and attempt <= max_retries:
                    sleep = base_sleep * (2 ** (attempt - 1)) * (0.8 + 0.4 * random.random())
                    logging.warning("API %s: %s, retrying in %.1fs (attempt %d/%d)",
                                    p.get("action"), code, sleep, attempt, max_retries)
                    time.sleep(sleep)
                    continue
                if code in RETRYABLE_AUTH_CODES and attempt <= max_retries:
                    logging.warning("API %s: %s — refreshing CSRF and retrying (attempt %d/%d)",
                                    p.get("action"), code, attempt, max_retries)
                    token = fetch_csrf(session, api_url, user_agent)
                    continue

                # Non-retryable error
                raise RuntimeError(f"API error for {p.get('action')}: {j['error']}")

            return j, token

        except requests.RequestException as e:
            # network/transport retry
            if attempt <= max_retries:
                sleep = base_sleep * (2 ** (attempt - 1)) * (0.8 + 0.4 * random.random())
                logging.warning("Transport error on %s: %s — retrying in %.1fs (attempt %d/%d)",
                                p.get("action"), e, sleep, attempt, max_retries)
                time.sleep(sleep)
                continue
            raise


# ====================================================================
#  Functions: SDC — statement inspection & mutation (P7482 with qualifiers)
# Everything needed to ensure the “source of file” (P7482) claim exists with the right
# value and qualifiers, add the Delpher resolver (P973) and operator (P137), and
# check/avoid duplication. ensure_statement orchestrates the create/update flow and
# bubbles up a refreshed CSRF token when needed.
# ====================================================================
def get_p7482_claims(session: requests.Session, mid: str) -> List[Dict]:
    """
    Return all existing **P7482** (“source of file”) claims for a given MediaInfo entity.

    Behavior
    --------
    - Validates the `mid` shape (`M` + digits). If invalid, logs an error and returns `[]`.
    - Calls `action=wbgetclaims&entity=<mid>&property=P7482`.
    - If the entity has **no SDC** (the API may respond with `no-such-entity`), returns `[]`.
    - For any other HTTP error, re-raises the exception (so callers can handle/abort).
    - If the JSON payload is missing or oddly shaped (no `"claims"` dict or no `P7482` key),
      logs a warning and returns `[]`.

    Args:
        session: Authenticated `requests.Session`.
        mid: MediaInfo ID like `"M109018409"` (case-insensitive ‘m’ also accepted).

    Returns:
        A list of P7482 claim objects (possibly empty). Each item is a standard
        Wikibase claim dict as returned by the Action API.

    Raises:
        requests.HTTPError: For HTTP failures **other than** the handled `no-such-entity`.
        TypeError: If inputs are of the wrong type.
    """
    if not isinstance(session, requests.Session):
        raise TypeError("get_p7482_claims: 'session' must be a requests.Session")
    if not isinstance(mid, str) or not re.fullmatch(r"[Mm]\d+", mid.strip()):
        logging.error("get_p7482_claims: invalid MediaInfo ID: %r", mid)
        return []

    try:
        resp = mw_request(session, "GET", {
            "action": "wbgetclaims",
            "entity": mid,
            "property": P_SOURCE_OF_FILE,
            "format": "json",
        })
    except requests.HTTPError as e:
        # Files with no SDC sometimes return 'no-such-entity'
        try:
            j = e.response.json()
            if (j.get("error", {}) or {}).get("code") == "no-such-entity":
                return []
        except Exception:
            pass
        raise

    claims = resp.get("claims")
    if not isinstance(claims, dict):
        logging.warning("get_p7482_claims: unexpected payload (no 'claims' dict) for %s: %r", mid, resp)
        return []

    prop_claims = claims.get(P_SOURCE_OF_FILE, [])
    if not isinstance(prop_claims, list):
        logging.warning("get_p7482_claims: 'claims[%s]' is not a list for %s: %r", P_SOURCE_OF_FILE, mid, prop_claims)
        return []

    return prop_claims


def claim_has_value_q(claim: Dict, numeric_qid: int) -> bool:
    """
    Return True if a claim’s mainsnak is a concrete entity value pointing to the
    given Q-id (by numeric id), otherwise False.

    This is defensive against incomplete/atypical structures:
      - Missing `mainsnak`, non-dict shapes, or `snaktype` != "value" → False
      - Non-entity datavalue types → False
      - Accepts either `value["numeric-id"]` or (fallback) `value["id"] == "Q<id>"`

    Args:
        claim: A single claim object as returned by wbgetclaims.
        numeric_qid: The numeric part of the target item id (e.g. 74228490 for Q74228490).

    Returns:
        bool: True if the claim’s value is the specified Q-id; otherwise False.
    """
    try:
        if not isinstance(claim, dict):
            return False
        snak = claim.get("mainsnak")
        if not isinstance(snak, dict):
            return False
        if snak.get("snaktype") != "value":
            return False

        dv = snak.get("datavalue")
        if not isinstance(dv, dict):
            return False
        if dv.get("type") != "wikibase-entityid":
            return False

        val = dv.get("value")
        if not isinstance(val, dict):
            return False

        # Prefer numeric-id if present
        if "numeric-id" in val:
            try:
                return int(val["numeric-id"]) == int(numeric_qid)
            except Exception:
                return False

        # Fallback to string id like "Q74228490"
        qid_str = val.get("id")
        if isinstance(qid_str, str) and qid_str.upper().startswith("Q"):
            try:
                return int(qid_str[1:]) == int(numeric_qid)
            except Exception:
                return False

        return False
    except Exception:
        return False


def claim_has_qualifier_url(claim: Dict, prop: str, url: str) -> bool:
    """
    Return True iff the claim has at least one qualifier `prop` whose datavalue is a
    string equal (after trimming whitespace) to `url`.

    Expected structure per qualifier:
        claim["qualifiers"][prop] -> list of snaks
        snak["datavalue"]["type"] == "string"
        snak["datavalue"]["value"] == "<url>"

    Defensive behavior:
    - If `claim` isn’t a dict, or `prop`/`url` aren’t strings → False.
    - Missing/ill-formed qualifier lists → False.
    - Malformed snaks (no datavalue, wrong type, non-string value) are ignored.

    Notes:
    - Comparison is exact and case-sensitive after `.strip()` on both sides.
      (No canonicalization beyond trimming; scheme/host casing and trailing slashes
       must match to return True.)
    """
    if not isinstance(claim, dict) or not isinstance(prop, str) or not isinstance(url, str):
        return False

    qualifiers = (claim.get("qualifiers") or {}).get(prop)
    if not isinstance(qualifiers, list):
        return False

    target = url.strip()
    for snak in qualifiers:
        try:
            dv = (snak or {}).get("datavalue")
            if not isinstance(dv, dict):
                continue
            if dv.get("type") != "string":
                continue
            val = dv.get("value")
            if isinstance(val, str) and val.strip() == target:
                return True
        except Exception:
            # Skip malformed qualifier entries and continue checking others
            continue
    return False



def claim_has_qualifier_q(claim: Dict, prop: str, numeric_qid: int) -> bool:
    """
    Return True iff the claim has at least one qualifier `prop` whose datavalue is a
    Wikibase entity pointing to the item with numeric-id == `numeric_qid`.

    Expected qualifier structure:
        claim["qualifiers"][prop] -> list of snaks
        snak["datavalue"]["type"] == "wikibase-entityid"
        snak["datavalue"]["value"]["numeric-id"] == <int>

    Defensive behavior:
    - If `claim` isn’t a dict, `prop` isn’t a str, or `numeric_qid` isn’t an int → False.
    - Missing/ill-formed qualifier lists → False.
    - Malformed snaks are ignored; the function scans all qualifiers.

    Notes:
    - Some APIs may include an `"id": "Q12345"` string instead of (or alongside) `"numeric-id"`.
      As a fallback, this function will also accept that form by parsing the digits and
      comparing to `numeric_qid`.
    """
    if not isinstance(claim, dict) or not isinstance(prop, str) or not isinstance(numeric_qid, int):
        return False

    qualifiers = (claim.get("qualifiers") or {}).get(prop)
    if not isinstance(qualifiers, list):
        return False

    for snak in qualifiers:
        try:
            dv = (snak or {}).get("datavalue")
            if not isinstance(dv, dict) or dv.get("type") != "wikibase-entityid":
                continue
            val = dv.get("value")
            if not isinstance(val, dict):
                continue

            # Preferred: numeric-id
            if val.get("numeric-id") == numeric_qid:
                return True

            # Fallback: "id": "Q12345"
            v_id = val.get("id")
            if isinstance(v_id, str) and v_id.startswith("Q"):
                try:
                    if int(v_id[1:]) == numeric_qid:
                        return True
                except Exception:
                    pass
        except Exception:
            # Skip malformed qualifier entries and continue checking others
            continue
    return False


def create_p7482_claim(session: requests.Session, csrf: str, mid: str) -> Tuple[str, str]:
    """
    Create a **P7482 = Q74228490** claim on a MediaInfo entity.

    This issues `action=wbcreateclaim` to add the main snak (P7482) with value
    “file available on the internet” (Q74228490). The call is routed through
    `mw_mutate`, which handles CSRF refresh and polite retry/backoff. If the CSRF
    token is refreshed during the mutation, the updated token is returned so callers
    can continue using the newest token.

    Args:
        session: Authenticated `requests.Session`.
        csrf: Current CSRF token string (may be refreshed by `mw_mutate`).
        mid: MediaInfo ID like `"M109018409"`.

    Returns:
        Tuple[str, str]: `(claim_id, csrf_token_used_or_refreshed)`, where `claim_id`
        is the newly created claim’s ID (e.g., `"M109018409$ABC-UUID"`).

    Raises:
        TypeError: If `session`, `csrf`, or `mid` have invalid types.
        ValueError: If `mid` is not of the form `M` + digits, or `csrf` is empty.
        RuntimeError: If the API responds with an `"error"` payload or the response
            is missing the expected `"claim"` structure.
        requests.RequestException: Propagated for transport-level failures.
    """
    # --- Input validation
    if not isinstance(session, requests.Session):
        raise TypeError("create_p7482_claim: 'session' must be a requests.Session")
    if not isinstance(csrf, str) or not csrf.strip():
        raise ValueError("create_p7482_claim: 'csrf' must be a non-empty string")
    if not isinstance(mid, str) or not re.fullmatch(r"[Mm]\d+", mid.strip()):
        raise ValueError(f"create_p7482_claim: invalid MediaInfo ID: {mid!r}")

    params = {"action": "wbcreateclaim"}
    value = json.dumps({"entity-type": "item", "numeric-id": Q_FILE_ONLINE})
    data = {
        "entity": mid,
        "property": P_SOURCE_OF_FILE,
        "snaktype": "value",
        "value": value,
    }

    j, new_token = mw_mutate(
        session,
        csrf,
        params,
        data,
        api_url=API_URL,
        user_agent=USER_AGENT,
    )

    # Validate expected structure
    claim = (j or {}).get("claim")
    if not isinstance(claim, dict) or "id" not in claim:
        raise RuntimeError(f"wbcreateclaim error: unexpected response {j!r}")

    return claim["id"], new_token

def add_qualifier_url(
    session: requests.Session,
    csrf: str,
    claim_id: str,
    prop: str,
    url: str
) -> Tuple[Dict, str]:
    """
    Add a URL qualifier to an existing Wikibase claim.

    This performs a POST `action=wbsetqualifier` with:
        snaktype = "value"
        value    = "<url>" (JSON-encoded string)

    CSRF handling:
        Uses the shared `mw_mutate(...)` helper (with retry/backoff + token refresh).
        The CSRF token may be refreshed; the updated token is returned.

    Parameters
    ----------
    session : requests.Session
        Authenticated session to the Commons Action API.
    csrf : str
        Current CSRF token (may be refreshed).
    claim_id : str
        Target claim id (e.g., "M123$ABC-UUID").
    prop : str
        Qualifier property id (e.g., "P973").
    url : str
        Absolute URL to add as the qualifier value (e.g., "https://resolver.kb.nl/resolve?urn=...").

    Returns
    -------
    Tuple[Dict, str]
        (resp_json, csrf) where `resp_json` is the API JSON response and `csrf`
        is the (possibly refreshed) CSRF token.
    Raises
    ------
    ValueError
        If inputs are malformed (empty claim_id/prop, or invalid URL).
    RuntimeError
        If the API returns an "error" payload.
    requests.HTTPError
        If the HTTP request fails (raised by `mw_mutate`).

    Notes
    -----
    - Caller should persist the returned CSRF for subsequent mutations in this run.
    - Includes `bot=1`, `assert=user`, and `maxlag=5` for politeness and identity guarantees.
    """
    # ---- Validate inputs (fail fast) ----
    if not isinstance(claim_id, str) or not claim_id.strip():
        raise ValueError(f"add_qualifier_url: invalid claim_id={claim_id!r}")
    if not isinstance(prop, str) or not prop.strip():
        raise ValueError(f"add_qualifier_url: invalid prop={prop!r}")
    if not isinstance(url, str) or not url.strip():
        raise ValueError("add_qualifier_url: url must be a non-empty string")

    url_str = url.strip()
    parsed = urlparse(url_str)
    if parsed.scheme not in ("http", "https") or not parsed.netloc:
        raise ValueError(f"add_qualifier_url: invalid absolute URL {url_str!r}")

    # ---- Build request ----
    params = {"action": "wbsetqualifier", "format": "json", "maxlag": "5"}
    data = {
        "claim": claim_id,
        "property": prop,
        "snaktype": "value",
        "value": json.dumps(url_str),  # value must be a JSON string
        "token": csrf,
        "bot": "1",
        "assert": "user",
    }

    # Use your mutate helper so badtoken/ratelimit are handled uniformly.
    resp_json, csrf = mw_mutate(session, csrf, params, data, api_url=API_URL, user_agent=USER_AGENT)

    if "error" in resp_json:
        raise RuntimeError(f"wbsetqualifier (url) error: {resp_json['error']}")

    return resp_json, csrf


def add_qualifier_q(
    session: requests.Session,
    csrf: str,
    claim_id: str,
    prop: str,
    numeric_qid: int
) -> Tuple[Dict, str]:
    """
    Add a Wikibase qualifier that points to an item (Q-id) to an existing claim.

    This performs a POST `action=wbsetqualifier` with:
        value = {"entity-type": "item", "numeric-id": <numeric_qid>}

    CSRF handling:
        Uses the shared `mw_mutate(...)` helper (retry/backoff + token refresh).
        The CSRF token may be refreshed; the updated token is returned to the caller.

    Parameters
    ----------
    session : requests.Session
        Authenticated session to the Commons Action API.
    csrf : str
        Current CSRF token (may be refreshed).
    claim_id : str
        Target claim id (e.g., "M123$ABC-UUID").
    prop : str
        Qualifier property id (e.g., "P973", "P137").
    numeric_qid : int
        Numeric item id (e.g., 20670235 for Q20670235).

    Returns
    -------
    Tuple[Dict, str]
        (resp_json, csrf) where `resp_json` is the API JSON response and `csrf`
        is the (possibly refreshed) CSRF token.

    Raises
    ------
    ValueError
        If inputs are malformed (empty claim_id/prop or non-positive numeric_qid).
    RuntimeError
        If the API returns an "error" payload.
    requests.HTTPError
        If the HTTP request fails (raised by `mw_mutate`).

    Notes
    -----
    - Caller should persist the returned CSRF for subsequent mutations in this run.
    - Includes `bot=1`, `assert=user`, and `maxlag=5` for politeness and identity guarantees.
    """
    # Basic input validation (fail fast)
    if not isinstance(claim_id, str) or not claim_id.strip():
        raise ValueError(f"add_qualifier_q: invalid claim_id={claim_id!r}")
    if not isinstance(prop, str) or not prop.strip():
        raise ValueError(f"add_qualifier_q: invalid prop={prop!r}")
    if not isinstance(numeric_qid, int) or numeric_qid <= 0:
        raise ValueError(f"add_qualifier_q: invalid numeric_qid={numeric_qid!r}")

    params = {"action": "wbsetqualifier", "format": "json", "maxlag": "5"}
    data = {
        "claim": claim_id,
        "property": prop,
        "snaktype": "value",
        "value": json.dumps({"entity-type": "item", "numeric-id": numeric_qid}),
        "token": csrf,
        "bot": "1",
        "assert": "user",
    }

    # Use your mutate helper so badtoken/ratelimit are handled uniformly.
    resp_json, csrf = mw_mutate(session, csrf, params, data, api_url=API_URL, user_agent=USER_AGENT)

    if "error" in resp_json:
        raise RuntimeError(f"wbsetqualifier (Q) error: {resp_json['error']}")

    return resp_json, csrf


def ensure_statement(session: requests.Session, csrf: str, mid: str, resolver_url: str) -> Tuple[str, Optional[str], str]:
    """
    Ensure that a MediaInfo entity has the statement **P7482 = Q74228490** (“file available on the internet”)
    qualified with **P973 = <resolver_url>** and **P137 = Q20670235 (Delpher)**.

    Behavior
    --------
    - Reads existing P7482 claims on the entity.
    - If no P7482=Q74228490 exists:
        * creates a new claim,
        * adds *both* qualifiers (P973=<resolver_url>, P137=Q20670235),
        * returns ("created", <claim_id>, <possibly refreshed csrf>).
    - If one exists but is missing either/both qualifiers:
        * adds the missing qualifier(s),
        * returns ("updated", <claim_id>, <possibly refreshed csrf>).
    - If everything is already present:
        * returns ("already-present", <claim_id>, <csrf unchanged>).

    CSRF Refresh
    ------------
    Helper functions (`create_p7482_claim`, `add_qualifier_url`, `add_qualifier_q`) may refresh the CSRF token
    (e.g., after a `badtoken` from the API). Therefore this function always returns the (possibly) updated `csrf`.

    Parameters
    ----------
    session : requests.Session
        An authenticated session to the Commons Action API.
    csrf : str
        A valid CSRF token. May be refreshed and returned updated.
    mid : str
        MediaInfo ID (e.g., "M109018409").
    resolver_url : str
        The Delpher resolver URL to store in qualifier P973.

    Returns
    -------
    Tuple[str, Optional[str], str]
        (status, claim_id_used, csrf), where:
          - status ∈ {"created", "updated", "already-present"}
          - claim_id_used is the claim id string (e.g., "M123$...") or None if creation somehow failed
          - csrf is the (possibly refreshed) CSRF token to be used by the caller for subsequent edits
    Raises
    ------
    ValueError
        If inputs are malformed (e.g., missing/invalid `mid` or `resolver_url`).
    requests.HTTPError, RuntimeError
        Propagated from lower-level API helpers with full context. Callers can catch and
        record per-row errors without aborting the entire batch.
    Notes
    -----
    - This function is intentionally strict about inputs but does not swallow network/API exceptions;
      callers typically wrap it in a try/except to set per-row Excel status.
    - Requires `assert=user`/`bot=1` and `maxlag` to be handled in the lower-level helpers.
    """
    # --- minimal input validation (fail fast with clear errors) ---
    if not isinstance(mid, str) or not mid.strip():
        raise ValueError(f"ensure_statement: invalid mid={mid!r}")
    if not isinstance(resolver_url, str) or not resolver_url.strip() or not resolver_url.startswith(("http://", "https://")):
        raise ValueError(f"ensure_statement: invalid resolver_url={resolver_url!r}")

    # Read existing claims (may raise HTTPError/RuntimeError → let caller handle)
    existing = get_p7482_claims(session, mid)

    # Find first P7482=Q74228490, if any
    target = next((cl for cl in existing if claim_has_value_q(cl, Q_FILE_ONLINE)), None)

    # Create new claim + both qualifiers
    if target is None:
        cid, csrf = create_p7482_claim(session, csrf, mid)  # returns (claim_id, possibly refreshed csrf)
        _, csrf = add_qualifier_url(session, csrf, cid, P_DESCRIBED_AT_URL, resolver_url)
        _, csrf = add_qualifier_q(session, csrf, cid, P_OPERATOR, Q_DELPHER)
        return "created", cid, csrf

    # Ensure qualifiers on existing claim
    need_update = False
    cid = target.get("id")

    if not claim_has_qualifier_url(target, P_DESCRIBED_AT_URL, resolver_url):
        _, csrf = add_qualifier_url(session, csrf, cid, P_DESCRIBED_AT_URL, resolver_url)
        need_update = True

    if not claim_has_qualifier_q(target, P_OPERATOR, Q_DELPHER):
        _, csrf = add_qualifier_q(session, csrf, cid, P_OPERATOR, Q_DELPHER)
        need_update = True

    return ("updated" if need_update else "already-present"), cid, csrf


# ==================================
# Functions: SDC — labels / captions
# Read existing captions, fetch the canonical
# File title (correct diacritics), derive a clean caption, and set
# missing labels in en/nl without overwriting existing ones.
# ==================================

def get_existing_labels(session: requests.Session, mid: str) -> Dict[str, str]:
    """
    Fetch existing MediaInfo labels (captions) for an entity and return a simple
    `{language: value}` mapping.

    Behavior:
        - Performs a read-only Action API call (`wbgetentities`) requesting `labels`.
        - If the MediaInfo entity does not exist (e.g., file has no SDC yet),
          returns an empty dict.
        - Filters out empty/whitespace-only label values.

    Args:
        session: An authenticated `requests.Session` for Commons.
        mid (str): MediaInfo ID like `"M109018409"` (case-insensitive `m` allowed).
    Returns:
        Dict[str, str]: A mapping of language codes to non-empty label strings.
                        Example: `{"en": "Rotterdamsche courant 21-11-1840", "nl": "…"}`.
                        Returns `{}` if there are no labels or no SDC entity.
    Raises:
        ValueError: If `mid` is not a valid MediaInfo ID.
        requests.HTTPError: For HTTP-layer failures other than `no-such-entity`.
        RuntimeError: If the API returns an unexpected structure.
    Notes:
        - This function is intentionally tolerant of “no-such-entity” and treats
          it as “no labels yet.” Callers can safely proceed to create labels.
        - The response may normalize the entity ID; we access the requested key
          and, if missing, fall back to the first entity in the payload.
    """
    # Validate MID shape and normalize to "M<digits>"
    if not isinstance(mid, str) or not mid.strip():
        raise ValueError(f"get_existing_labels: empty or invalid mid: {mid!r}")
    m = re.fullmatch(r"[Mm](\d+)", mid.strip())
    if not m:
        raise ValueError(f"get_existing_labels: MID must be 'M<digits>', got {mid!r}")
    mid_norm = f"M{m.group(1)}"

    # Call API
    try:
        resp = mw_request(session, "GET", {
            "action": "wbgetentities",
            "ids": mid_norm,
            "props": "labels",
            "format": "json",
        })
    except requests.HTTPError as e:
        # Treat “no-such-entity” as "no labels yet"
        try:
            j = e.response.json() if e.response is not None else {}
            if (j.get("error") or {}).get("code") == "no-such-entity":
                return {}
        except Exception:
            pass
        raise

    if not isinstance(resp, dict):
        raise RuntimeError(f"wbgetentities: unexpected response type: {type(resp).__name__}")

    entities = resp.get("entities")
    if not isinstance(entities, dict) or not entities:
        # No entities key (or empty) → nothing to return
        return {}

    ent = entities.get(mid_norm)
    if ent is None:
        # Fallback: sometimes the API returns only one entity under a normalized key
        ent = next(iter(entities.values()), {})

    labels = ent.get("labels") or {}
    if not isinstance(labels, dict):
        return {}

    out: Dict[str, str] = {}
    for lang, obj in labels.items():
        val = (obj or {}).get("value", "")
        if isinstance(val, str):
            s = val.strip()
            if s:
                out[lang] = s
    return out


def get_live_file_title(session: requests.Session, pageid: int) -> Optional[str]:
    """
    Fetch the canonical File: title for a given page ID (with correct diacritics).

    Behavior:
        - Performs a read-only Action API `query` for `prop=info` on the given `pageid`.
        - Returns the page title string (e.g., "File:Example.pdf") if present.
        - Returns `None` if the page array is empty or the page is marked as missing.
    Args:
        session: An authenticated `requests.Session` configured for Commons.
        pageid: Positive integer MediaWiki page ID (namespace 6 / File:).
    Returns:
        Optional[str]: The canonical title, or `None` if not available.
    Raises:
        ValueError: If `pageid` is not a positive integer.
        requests.HTTPError: Propagated for HTTP/API transport errors (so the caller
            can log HTTP status and error codes).
        RuntimeError: If the API returns an unexpected structure.
    Notes:
        - This function is intentionally tolerant of “missing page” and simply
          returns `None` in that case. Callers can decide how to handle it.
        - We do not swallow `requests.HTTPError`; letting it bubble up allows the
          caller’s try/except to record precise HTTP context.
    """
    # Validate input early
    if not isinstance(pageid, int) or pageid <= 0:
        raise ValueError(f"get_live_file_title: pageid must be a positive int, got {pageid!r}")

    # API call (let HTTP errors propagate to caller)
    resp = mw_request(session, "GET", {
        "action": "query",
        "format": "json",
        "formatversion": "2",
        "prop": "info",
        "pageids": str(pageid),
    })

    # Parse and validate structure
    if not isinstance(resp, dict):
        raise RuntimeError(f"get_live_file_title: unexpected response type {type(resp).__name__}")

    pages = (resp.get("query") or {}).get("pages")
    if not isinstance(pages, list) or not pages:
        return None

    page = pages[0] or {}
    if page.get("missing"):
        return None

    title = page.get("title")
    return title if isinstance(title, str) and title.strip() else None

def set_label(session, csrf, mid, lang, value):
    """
    Set a MediaInfo label (caption) via the Action API (`wbsetlabel`).

    This function performs a single, strict call to `wbsetlabel` to set the label for
    the given MediaInfo entity **without checking** if a label already exists. If you
    want to avoid overwriting existing captions, make sure the caller checks current
    labels first (e.g., with `wbgetentities`) and only calls this setter when missing.

    Args:
        session: An authenticated `requests.Session` (post-login) for Commons.
        csrf (str): A valid CSRF token string. (May be refreshed by `mw_mutate`.)
        mid (str): MediaInfo ID, e.g. "M109018409".
        lang (str): Target language code, e.g. "en" or "nl".
        value (str): The caption text to set (non-empty).
    Returns:
        dict: The JSON response from the API on success (e.g., `{"success":1, ...}`).
    Raises:
        ValueError: If inputs are missing/invalid (empty strings, wrong MID shape, etc.).
        RuntimeError: If the API returns an error or the request/transport fails.
    Notes:
        - This uses a helper `mw_mutate(session, csrf, params, data, ...)` which is expected
          to add required fields like `token`, `assert=user`, `bot=1`, `maxlag`, and handle
          a single token refresh if needed. If `mw_mutate` returns a new CSRF token, it is
          ignored here; the caller can capture it if desired.
        - The call is intentionally strict: any API error results in a `RuntimeError` so
          the caller can catch it and continue per-row without crashing the whole batch.
    """
    # Basic validation (fail fast, clear messages)
    if not isinstance(mid, str) or not mid.strip() or not re.fullmatch(r"[Mm]\d+", mid.strip()):
        raise ValueError(f"set_label: invalid MediaInfo ID: {mid!r}")
    if not isinstance(lang, str) or not lang.strip():
        raise ValueError("set_label: 'lang' must be a non-empty string.")
    if not isinstance(value, str) or not value.strip():
        raise ValueError("set_label: 'value' must be a non-empty string.")

    params = {"action": "wbsetlabel"}
    data = {"id": mid.strip(), "language": lang.strip(), "value": value}

    try:
        j, _new_csrf = mw_mutate(session, csrf, params, data, api_url=API_URL, user_agent=USER_AGENT)
    except requests.HTTPError as he:
        code = getattr(getattr(he, "response", None), "status_code", "?")
        try:
            errj = he.response.json() if he.response is not None else {}
            err_code = (errj.get("error") or {}).get("code", "")
        except Exception:
            err_code = ""
        raise RuntimeError(f"wbsetlabel HTTP {code} {err_code}") from he
    except Exception as e:
        raise RuntimeError(f"wbsetlabel transport/error: {e}") from e

    if isinstance(j, dict) and "error" in j:
        err = j["error"]
        raise RuntimeError(f"wbsetlabel API error: {err.get('code')} - {err.get('info')}")

    return j


# ==================================
# Functions: Orchestration main flow
# The end-to-end pipeline: load Excel, slice workload, login, iterate
# rows with logging/progress, ensure statements, add captions, checkpoint-write
# to Excel periodically, and final write-out.
# ==================================

def main() -> None:
    """
    Run one end-to-end SDC update pass driven by the Excel input.

    High-level flow
    ---------------
    1) Load and validate the Excel workbook/sheet (EXCEL_FILE / EXCEL_SHEET).
       - Require the columns defined by EXCEL_URL_COL and EXCEL_RESOLVER_COL.
       - Ensure the status/candidate columns exist and are typed as pandas StringDtype:
           * EXCEL_STATUS_COL
           * EXCEL_CAPTION_EN_STATUS
           * EXCEL_CAPTION_NL_STATUS
           * EXCEL_CAPTION_CANDIDATE
         (Existing columns are coerced to StringDtype and NA values are filled with "").

    2) Resolve the real sheet *name* (useful when EXCEL_SHEET is an index) so that a later
       write-back replaces the same sheet.

    3) Optionally slice the workload with apply_slice(df) using HEAD or RANGE (mutually
       exclusive, strict validation). The returned DataFrame ‘work’ is what gets processed.

    4) Authenticate to the MediaWiki Action API and obtain a CSRF token via
       mw_login_and_tokens().

    5) Set up tqdm-aware logging (setup_tqdm_logging) and iterate rows with a progress bar
       that displays “1/total” while the first row is being handled.

    Per-row behavior
    ----------------
    Given the row index `i` (0-based) and record:
    - Extract the MediaInfo ID `mid` from the URL (extract_mid). If absent:
      * Log an error, set EXCEL_STATUS_COL to a descriptive “error: …” string, and mark both
        caption status columns as “skipped-empty-candidate” (if still empty). Continue to next row.

    - Build a clickable Commons entity URL for visibility in logs:
        https://commons.wikimedia.org/entity/{mid}

    - Derive the numeric `pageid` from `mid` locally (pageid_from_mid) and fetch the live File
      title by pageid (get_live_file_title). The live title (with correct diacritics) is used
      to derive the caption candidate.

    - Statements: ensure P7482 = Q74228490 (file available on the internet) exists for `mid`
      with qualifiers:
        * P973 (described at URL) = ResolverURL from Excel row
        * P137 (operator) = Q20670235 (Delpher)
      This is done via ensure_statement(session, csrf, mid, resolver), which may refresh the
      CSRF token internally and returns a status in {"created","updated","already-present"}.
      The status is written to EXCEL_STATUS_COL. Any HTTP/API error is logged and recorded
      as “error:<message>” without aborting the run.

    - Captions: derive a caption candidate from the *live* title (derive_caption_from_live_title)
      and write it to EXCEL_CAPTION_CANDIDATE for transparency/QA. If a non-empty candidate is
      available:
        * Read existing labels (get_existing_labels).
        * If “en” is missing, set it via set_label(…, "en", candidate) and mark
          EXCEL_CAPTION_EN_STATUS="created"; otherwise “already-present”.
        * Ditto for “nl”.
      If the candidate is empty or the pageid/title cannot be resolved, the caption status
      columns are set to “skipped-empty-candidate” or “error:<message>” as appropriate.

    - Optional convenience: if OPEN_AFTER_SUCCESS is enabled and a caption was created, open
      the entity page in the browser (up to OPEN_AFTER_SUCCESS_MAX tabs per run).

    - Checkpointing: if this row performed any change (statement created/updated or caption created),
      increment a counter; when it reaches CHECKPOINT_EVERY_SUCCESS, write the entire DataFrame
      back to the same sheet (write_back) and reset the counter. This ensures progress is saved
      periodically, even if a run is interrupted later.

    - Politeness: sleep SLEEP_BETWEEN seconds between rows; all mutating requests set maxlag=5.

    Error handling
    --------------
    - Fatal configuration/read errors (missing Excel, bad sheet name/index, missing required
      columns) log at CRITICAL and exit via SystemExit(1).
    - Per-row API/HTTP errors are caught, logged with context (row number, MID, entity URL),
      and recorded into the Excel status columns; the loop continues with the next row.
    - CSRF token refreshes are handled inside ensure_statement / set_label (via mw_mutate),
      so the main loop does not need to special-case token expiry.

    Side effects
    ------------
    - Network calls to commons.wikimedia.org for both read and write operations.
    - Periodic writes of the updated DataFrame back to EXCEL_FILE (same sheet name).
    - Optional opening of browser tabs.
    - Console logging interleaved cleanly with the tqdm progress bar.

    Environment
    -----------
    Uses the following configuration (see module constants for defaults):
    - EXCEL_FILE, EXCEL_SHEET, EXCEL_URL_COL, EXCEL_RESOLVER_COL
    - EXCEL_STATUS_COL, EXCEL_CAPTION_EN_STATUS, EXCEL_CAPTION_NL_STATUS, EXCEL_CAPTION_CANDIDATE
    - HEAD, RANGE, SLEEP_BETWEEN, OPEN_AFTER_SUCCESS, OPEN_AFTER_SUCCESS_MAX
    - CHECKPOINT_EVERY_SUCCESS
    - WIKIMEDIA_USERNAME, WIKIMEDIA_PASSWORD, WIKIMEDIA_USER_AGENT

    Returns
    -------
    None

    Raises
    ------
    SystemExit
        On unrecoverable configuration or Excel I/O errors (see “Error handling” above).
    """

    # Read Excel
    if not os.path.exists(EXCEL_FILE):
        logging.critical("Excel not found: %s", EXCEL_FILE); raise SystemExit(1)
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=EXCEL_SHEET)
    except Exception as e:
        logging.critical("Failed reading Excel '%s' (sheet=%r): %s", EXCEL_FILE, EXCEL_SHEET, e); raise SystemExit(1)

    # Require URL + ResolverURL columns
    for col in (EXCEL_URL_COL, EXCEL_RESOLVER_COL):
        if col not in df.columns:
            logging.critical("Missing required column %r. Available: %s", col, list(df.columns)); raise SystemExit(1)

    # Ensure status/candidate columns exist (create if missing)
    for col in (EXCEL_STATUS_COL, EXCEL_CAPTION_EN_STATUS, EXCEL_CAPTION_NL_STATUS, EXCEL_CAPTION_CANDIDATE):
        if col not in df.columns:
            # create as proper string dtype from the start
            df[col] = pd.Series(pd.array([""] * len(df), dtype="string"))
        else:
            # coerce existing column to pandas StringDtype and replace NaN with empty string
            df[col] = df[col].astype("string").fillna("")

    # Resolve a proper sheet name for writing
    try:
        wb = load_workbook(EXCEL_FILE, read_only=True); sheetnames = wb.sheetnames; wb.close()
    except Exception as e:
        logging.critical("Cannot read sheet names: %s", e); raise SystemExit(1)
    if isinstance(EXCEL_SHEET, int):
        if 0 <= EXCEL_SHEET < len(sheetnames):
            target_sheet = sheetnames[EXCEL_SHEET]
        else:
            logging.critical("Sheet index %r out of range. Available: %s", EXCEL_SHEET, sheetnames); raise SystemExit(1)
    elif isinstance(EXCEL_SHEET, str) and EXCEL_SHEET.strip():
        target_sheet = EXCEL_SHEET
    else:
        target_sheet = sheetnames[0]

    # Slice the workload (HEAD/RANGE)
    work = apply_slice(df)

    # Login (Action API session + CSRF)
    session, csrf = mw_login_and_tokens()

    opened = 0
    success_since_flush = 0  # count of rows where we changed something (statement created/updated or caption created)

     # Progress bar that displays 1/total while processing the first row
    setup_tqdm_logging()

    total_rows = len(work)
    with tqdm(total=len(work), desc="SDC updates", unit="row", dynamic_ncols=True, leave=True, mininterval=0) as pbar:
        for idx, (i, row) in enumerate(work.iterrows(), start=1):
            # Visually show "idx/total" *before* doing the work on this row
            pbar.n = idx
            pbar.refresh()
            url = str(row.get(EXCEL_URL_COL) or "")
            resolver = str(row.get(EXCEL_RESOLVER_COL) or "")
            mid = extract_mid(url)
            if not mid:
                bad = (url or "").strip()
                msg = f"error: no MID in URL {bad!r}"
                logging.error("Row %d: %s — %s", i + 1, msg, bad)
                df.at[i, EXCEL_STATUS_COL] = msg
                # captions can’t be derived either; mark as skipped if still empty
                if not df.at[i, EXCEL_CAPTION_EN_STATUS]:
                    df.at[i, EXCEL_CAPTION_EN_STATUS] = "skipped-empty-candidate"
                if not df.at[i, EXCEL_CAPTION_NL_STATUS]:
                    df.at[i, EXCEL_CAPTION_NL_STATUS] = "skipped-empty-candidate"
                # continue to next row (bar already reflects idx)
                continue

            # Pick one (entity page is nice for clicking in logs)
            entity_url = f"https://commons.wikimedia.org/entity/{mid}"
            # (If you prefer the JSON view for debugging, use:)
            # entity_url = f"https://commons.wikimedia.org/wiki/Special:EntityData/{mid}"

            # Resolve pageid & LIVE title (with a single quick retry if needed)
            pageid = pageid_from_mid(mid)
            if pageid is None:
                time.sleep(1.0)
                pageid = pageid_from_mid(mid)
            live_title = get_live_file_title(session, pageid) if pageid else None

            row_had_change = False  # track whether this row performed any change

            # --- Ensure the statement with qualifiers (if resolver looks OK) ---
            try:
                if resolver and resolver.startswith("http"):
                    status, _cid, csrf = ensure_statement(session, csrf, mid, resolver)
                    df.at[i, EXCEL_STATUS_COL] = status

                    logging.info("Row %d (%s): statement %s — %s", i+1, mid, status, entity_url)
                    if status in ("created", "updated"):
                        row_had_change = True
                else:
                    msg = f"error: bad ResolverURL {resolver!r}"
                    logging.error("Row %d (%s): %s", i+1, mid, msg)
                    df.at[i, EXCEL_STATUS_COL] = msg
            except requests.HTTPError as he:
                try:
                    errj = he.response.json()
                    em = f"HTTP {he.response.status_code} {errj.get('error', {}).get('code','')}"
                except Exception:
                    em = f"HTTP {getattr(he.response,'status_code','?')}"
                logging.error("Row %d (%s): %s — %s", i+1, mid, em, entity_url)
                df.at[i, EXCEL_STATUS_COL] = f"error:{em}"
            except Exception as e:
                logging.error("Row %d (%s): statement error: %s — %s", i+1, mid, e, entity_url)
                df.at[i, EXCEL_STATUS_COL] = f"error:{e}"

            # --- Captions from LIVE title (do not overwrite existing) ---
            try:
                if not pageid:
                    raise RuntimeError("no pageid via Special:EntityData")
                if not live_title:
                    live_title = get_live_file_title(session, pageid)
                    if not live_title:
                        raise RuntimeError("no live title for pageid")

                candidate = derive_caption_from_live_title(live_title)
                df.at[i, EXCEL_CAPTION_CANDIDATE] = candidate

                if not candidate.strip():
                    df.at[i, EXCEL_CAPTION_EN_STATUS] = df.at[i, EXCEL_CAPTION_EN_STATUS] or "skipped-empty-candidate"
                    df.at[i, EXCEL_CAPTION_NL_STATUS] = df.at[i, EXCEL_CAPTION_NL_STATUS] or "skipped-empty-candidate"
                else:
                    labels = get_existing_labels(session, mid)

                    # EN
                    if "en" not in labels:
                        set_label(session, csrf, mid, "en", candidate)
                        df.at[i, EXCEL_CAPTION_EN_STATUS] = "created"
                        row_had_change = True
                    else:
                        df.at[i, EXCEL_CAPTION_EN_STATUS] = "already-present"

                    # NL
                    if "nl" not in labels:
                        set_label(session, csrf, mid, "nl", candidate)
                        df.at[i, EXCEL_CAPTION_NL_STATUS] = "created"
                        row_had_change = True
                    else:
                        df.at[i, EXCEL_CAPTION_NL_STATUS] = "already-present"

                    # Optionally open entity page after any caption change
                    if OPEN_AFTER_SUCCESS and opened < OPEN_AFTER_SUCCESS_MAX and (
                        df.at[i, EXCEL_CAPTION_EN_STATUS] == "created" or df.at[i, EXCEL_CAPTION_NL_STATUS] == "created"
                    ):
                        open_page_in_browser(mid, live_title)
                        opened += 1

                logging.info("Row %d (%s): captions checked/updated — %s", i+1, mid, entity_url)

            except requests.HTTPError as he:
                try:
                    errj = he.response.json()
                    em = f"HTTP {he.response.status_code} {errj.get('error', {}).get('code','')}"
                except Exception:
                    em = f"HTTP {getattr(he.response,'status_code','?')}"
                logging.error("Row %d (%s): captions error: %s — %s", i+1, mid, em, entity_url)
                if not df.at[i, EXCEL_CAPTION_EN_STATUS]:
                    df.at[i, EXCEL_CAPTION_EN_STATUS] = f"error:{em}"
                if not df.at[i, EXCEL_CAPTION_NL_STATUS]:
                    df.at[i, EXCEL_CAPTION_NL_STATUS] = f"error:{em}"

            except Exception as e:
                logging.error("Row %d (%s): captions error: %s — %s", i+1, mid, e, entity_url)
                if not df.at[i, EXCEL_CAPTION_EN_STATUS]:
                    df.at[i, EXCEL_CAPTION_EN_STATUS] = f"error:{e}"
                if not df.at[i, EXCEL_CAPTION_NL_STATUS]:
                    df.at[i, EXCEL_CAPTION_NL_STATUS] = f"error:{e}"

            # Checkpoint write after N successful updates (any change this row)
            if row_had_change:
                success_since_flush += 1
                if success_since_flush >= CHECKPOINT_EVERY_SUCCESS:
                    try:
                        write_back(df, target_sheet)
                        logging.info("Checkpoint: wrote statuses after %d successful updates.", success_since_flush)
                    except Exception as e:
                        logging.error("Checkpoint write failed: %s", e)
                    success_since_flush = 0

            # Polite delay
            if SLEEP_BETWEEN > 0:
                time.sleep(SLEEP_BETWEEN)

    # Final write-back
    write_back(df, target_sheet)
    logging.info("Done.")


if __name__ == "__main__":
    main()